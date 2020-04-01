# Team Biodigestables
# import modules
import datetime
import os
import PySimpleGUI as sg # allows creation of user interface
import pandas as pd # allows working with Excel files
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from pathlib import Path

def analyze(fileName,flow):
    
    dates, timeData, t1, t2, t3, t4 = [],[],[], [], [], []
    
    # open data file
    with open(fileName) as tsv:

        line = tsv.readline()
        
        while line:
            line = tsv.readline()         
            line = line.split(" | ")
            line.pop() # takes off new line char
            print(line)
            for i in range(len(line)):
                if i == 0:
                    dates.append(line[i])       
                elif i == 1:
                    timeData.append(line[i])
                elif i == 2:
                    t1.append(float(line[i]))
                elif i == 3:
                    t2.append(float(line[i]))
                elif i == 4:
                    t3.append(float(line[i]))
                elif i == 5:
                    t4.append(float(line[i]))
            
        print("i0, Dates", dates)
        print("i1, Times",timeData)
        print("i2, T1", t1)
        print("i3, T2", t2)
        print("i4, T3", t3)
                
    
    times = []
    
    for i in range(len(timeData)):   
        
        # convert time data to datetime
        x = datetime.datetime.strptime(timeData[i], '%H:%M:%S')
        times.append(x.time())

    #fileData = [dates,timeData,temp1Data,temp2Data,tempAmbData]
    fileData = [times,t1,t2,t3,t4]

    # find time it takes the tank to reach 30 C
    for i in range(len(t1)):
        timeAt30 = 'N/A'
        
        if t1[i] >= 30:
            timeAt30 = times[i]
            break
        
    # DEPENDENT VARIABLES
    # TIME CALCS
    startTime = datetime.datetime.combine(datetime.date.today(), times[0])
    print(startTime)
    endTime = datetime.datetime.combine(datetime.date.today(), times[-1])
    totalTime = endTime - startTime
    
    totalTime = [round(totalTime.total_seconds() / 3600,3)] # convert to hours
    
    # Get the difference between datetimes (as timedelta)
    if timeAt30 != 'N/A':
        timeAt30 = datetime.datetime.combine(datetime.date.today(), timeAt30)
        timeTo30 = timeAt30 - startTime
        timeTo30 = [round(timeTo30.total_seconds() / 3600,3)]
    else:
        timeTo30 = 'N/A'
    
    # MAX TEMP CALCS
    maxTemp1 = max(t1)
    maxIdx = t1.index(max(t1))
    maxTemp2 = max(t2) # max temp of control
    
    # MAX TIME CALCS - get time it takes to get to max temp
    maxTime = datetime.datetime.combine(datetime.date.today(), times[maxIdx])
    maxTime = maxTime - startTime
    maxTime = [round(maxTime.total_seconds() / 3600,3)]

    
    # find heat gain required to raise temperature to max temp
    cp = 4.184 # [kJ/kg*K] heat capacity of water
    heatGain = round(float(flow)*cp*(maxTemp1 - t1[0]),3)
    
    # get temp difference between initial temp and max temp
    tempDiff = round(maxTemp1 - t1[0],2)
    
    # getlow/high ambient temp
    minTempAmb = min(t3)
    maxTempAmb = max(t3)
    
    print("Max Temp 1: ", maxTemp1)
    print("Max Temp 2: ", maxTemp2)
    print("Time to Max: ", maxTime)
    print("Time to Reach 30 C: ", timeTo30)
    print("Temp Diff: ", tempDiff)
    print("Min/Max Temp Amb: ", minTempAmb, maxTempAmb)
    
    print("i0, Dates", dates)
    print("i1, Times",timeData)
    print("i2, T1", t1)
    print("i3, T2", t2)
    print("i4, T3", t3)
    print("i5, T4", t4)
    
    print(len(times),len(t1),len(t2),len(t3))
    print("Time vs. Tank Temp")
    plt.plot(times,t1)
    plt.show()
    print("Time vs. Ambient Temp")
    plt.plot(times,t2)
    plt.show()
    
    return [startTime],tempDiff,maxTemp1,maxTemp2,timeTo30,totalTime,maxTime,heatGain,minTempAmb,maxTempAmb,fileData


# WRITING TO EXCEL
def write(weather,flow,wind,startTime,totalTime,other,tempDiff,maxTemp1,maxTemp2,timeTo30,maxTime,heatGain,minTempAmb,maxTempAmb,fileName,fileData):
    
    if len(weather) == 0:
        weather = ['']
        
    fileName = Path(fileName).stem
    # reads the column headers from the Excel sheet
    df = pd.read_excel('solarData.xlsx',sheet_name='Sheet1')
    
    # These are the input headers in the data entry file
    headers = ['Testing Date', 'Weather Conditions', 'Avg Flow Rate [kg/s]', 'Wind Speed [m/s]', 'Start Time', 'Test Duration [hrs]', 'Other Tested Variables', 'Max Temp Diff [C]', 'Max Temp [C]', 'Time to Reach 30C [hrs]', 'Time to Reach Max Temp [hrs]', 'Heat Gain to Reach Max Temp [kW]']
    
    testData = {'Testing Date': [fileName],
                'Weather Conditions': weather,
                'Avg Flow Rate [kg/s]': [flow],
                'Wind Speed [m/s]': [wind],
                'Start Time': startTime,
                'Test Duration [hrs]': totalTime,
                'Min T_a [C]': [minTempAmb],
                'Max T_a [C]': [maxTempAmb],
                'Significant Test Notes: ':[other],
                'Max Control Temp [C]':[maxTemp2],
                'Max Tank Temp [C]': [maxTemp1],
                'Max Temp Diff [C]': [tempDiff],
                'Time to Reach 30C [hrs]': [timeTo30],
                'Time to Reach Max Temp [hrs]': maxTime,
                'Heat Gain to Reach Max Temp [kW]': [heatGain]}
    
    df = df.append(testData, ignore_index=True)
    print(testData.values())
    print("works")
    
    values = list(testData.values())
    for i in range(len(values)):
        print(values[i], ":")
        print(len(values[i]))
    # new dataframe with same columns
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter('solarData.xlsx', engine='openpyxl')
    writer.book = load_workbook('solarData.xlsx')
    
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    print("works")
    # read existing file
    reader = pd.read_excel(r'solarData.xlsx')
    df.to_excel(writer, sheet_name='Sheet1',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()

    # WRITING TO SECOND SHEET
    df = pd.read_excel('solarData.xlsx',sheet_name='Sheet2')
    testData = {'Date':fileName,'Time': fileData[0], 'Temp1 [C]': fileData[1], 'Temp2 [C]': fileData[2], 'Temp_a[C]': fileData[3]}
    print(fileData)
    print(fileName)
    df = df.append(testData, ignore_index=True,sort=False)
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter('solarData.xlsx', engine='openpyxl')
    writer.book = load_workbook('solarData.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(r'solarData.xlsx',sheet_name='Sheet2')
    df.to_excel(writer, sheet_name='Sheet2',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()
    
    
################
################
### GUI CODE ###
# creates the user interface
def_font = 'Verdana 10' # default font

# utility description
introMain = "Select the input temperature data file from your computer. \nFill in all the required independent variable fields and click \
'Analyze Data'. \n* = required field"

# adjusts the size, content, and spatial arrangements of the window elements
layout = [[sg.Text('HPD Biodigester Solar Model', font='Verdana 13 bold')],[sg.Text(introMain, font=def_font)],
			[sg.Text('Select Input Data File*:', font=def_font, size=(20, 1))],
            [sg.InputText(size=(50, 1), font=def_font, key='fileName'),sg.FileBrowse(file_types=(("Text Files", "*.txt*"),),tooltip='Selects a data file from your computer.')],
            
            [sg.Text('Enter Environment Variables:', font='Verdana 13 bold')],
            [sg.Text('Weather: '),sg.Listbox(values=['Clear and Sunny', 'Cloudy Sunny', 'Cloudy','Light Drizzle','Rain','Heavy Rain','Sleet','Snow'], size=(20, 4),key='weather')],
            [sg.Text('Wind Speed:'),sg.InputText('',size=(5, 1), font=def_font, key='wind'),sg.Text('m/s')],
            
            [sg.Text('Enter Independent Variables:', font='Verdana 13 bold')],
            [sg.Text('Avg Flow Rate*:',font=def_font, ),sg.InputText('',size=(10, 1), font=def_font, key='flow'),sg.Text('kg/s')],
            [sg.Text('Test Notes and/or Other Variables (if any): ')],
            [sg.Multiline('',size=(30, 3), font=def_font, key='other'),],
            
            [sg.Button('Analyze Data', font=def_font, tooltip='Initial Data Analysis'),],
            [sg.Text('Analysis Output: ',font='Verdana 13 bold')],
            [sg.Multiline('',size=(50,8),disabled=True,font=def_font, key='output')],
            [sg.Text('Click Update Master Data to save your analysis data to the "solarData.xlsx" Excel file.',key='update')],
            
			[sg.Button('Update Master Data', font=def_font, tooltip='Updates Master Excel sheet'),
            sg.Button('Clear', font=def_font, tooltip='Clears all fields.'),]]

window = sg.Window('Solar Model', layout, resizable=True)
while True:
    event, values = window.Read()
    
    # input vars
    fileName = values['fileName']
    weather = values['weather']
    flow = (values['flow'])
    wind = values['wind']
    other = values['other']
    
    if event is None:
        break
    
    if event == 'Analyze Data':
        
        # check to see if the user filled out required fields
        if values['flow'] == "":
            window.Element('update').Update("Please enter a flow rate.")
        elif values['fileName'] == "":
            window.Element('update').Update("Please select an input file.")
        else:
            startTime,tempDiff,maxTemp1,maxTemp2,timeTo30,totalTime,maxTime,heatGain,minTempAmb,maxTempAmb,fileData = analyze(fileName,flow)
            window.Element('output').Update("Outside Temp (Low/High): " + str(minTempAmb) + "/" + str(maxTempAmb) + " C\n"\
                                            + "Max Temp Diff: " + str(tempDiff) + " C\n"\
                                            + "Max Temp: " + str(maxTemp1) + " C\n"\
                                            + "Time to Reach 30C: " + str(timeTo30) + " hrs\n"\
                                            + "Time to Reach Max Temp: " + str(maxTime) + " hrs\n"\
                                            + "Heat Gained to Reach Max Temp: " + str(heatGain) + " kW\n")
    
    if event == 'Update Master Data':
        
        if len(values['output']) <= 1:
            window.Element('update').Update("Must Analyze Data before saving to Excel sheet.")
        
        else:
            # checks if the Excel file is open
            def not_in_use(filename):
                try:
                    # if file exists, check if it is open
                    if os.path.isfile(filename):
                        os.rename(filename,filename)
                        return True
                except:
                    return False
                    window.Element('update').Update("Please close 'solarData.xlsx' before updating master data.")
                
            write(weather,flow,wind,startTime,totalTime,other,tempDiff,maxTemp1,maxTemp2,timeTo30,maxTime,heatGain,minTempAmb,maxTempAmb,fileName,fileData)
            window.Element('update').Update("Your data has been successfully saved to 'solarData.xlsx'")
                
    
    # Clear button - clears all fields
    if event == 'Clear':
        window.Element('fileName').Update("")
        window.Element('wind').Update("")
        window.Element('other').Update("")
        window.Element('flow').Update("")
        window.Element('output').Update("")
        window.Element('update').Update("")
        
##########
##########

