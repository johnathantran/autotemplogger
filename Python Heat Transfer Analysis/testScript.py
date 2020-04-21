# Team Biodigester
# import modules
import datetime
import os
import pandas as pd # allows working with Excel files
from openpyxl import load_workbook
from pathlib import Path

class temperatureData:
    
    def __init__(self,**kwargs):
        self.dates = kwargs['dates']
        self.timeData = kwargs['times']
        
        if "t1" in kwargs:
            self.allTemps.append(kwargs['t1'])
            
    def analyzeData():
        return analysis
    
    # defines how an instance of the class is printed          
    def __str__(self):   
        printStr = ""       
        
        for date in self.dates:
            printStr += str(date)
            
        return printStr
    
        
def analyze(fileName,flow):

    dates, timeData, t1, t2, t3, t4 = [], [], [], [], [], []
    kwargs = {"dates": [], "times":[]}
 
    # open data file
    with open(fileName) as tsv:

        line = tsv.readline()

        while line:
            line = tsv.readline()
            
            # break the loop if the line is empty
            if line == '':
                break
            
            # converts line string to a list split on the "|" char
            line = [x.strip() for x in line.split('|')]
            line.pop() # takes off new line char
            print(line)
            

            for i, val in enumerate(line):
                    
                if i not in kwargs:
                    kwargs[str(i)] = []
                        
                if i == 0:
                    dates.append(val)
                    kwargs["dates"].append(val)
                    
                elif i == 1:
                    timeData.append(val)
                    kwargs["times"].append(val)
                    
                elif i == 2:
                    t1.append(float(val))      
                elif i == 3:
                    t2.append(float(val))
                elif i == 4:
                    t3.append(float(val))
                elif i == 5:
                    t4.append(float(val))
        
        
                if i > 1:
                    kwargs[str(i)].append(val)  
        
        
        
        print("kwargs: ", kwargs, "\n")
        instance = temperatureData(**kwargs)
        
        print("i0, Dates", dates)
        print("i1, Times",timeData)
        print("i2, T1", t1)
        print("i3, T2", t2)
        print("i4, T3", t3)


    times = []

    for i,val in enumerate(timeData):

        # convert time data to datetime
        x = datetime.datetime.strptime(val, '%H:%M:%S')
        times.append(x.time())

    #fileData = [dates,timeData,temp1Data,temp2Data,tempAmbData]
    fileData = [times,t1,t2,t3,t4]

    # find time it takes the tank to reach 30 C
    for i,val in enumerate(t1):
        timeAt30 = 'N/A'

        if val >= 30:
            timeAt30 = times[i]
            break

    # DEPENDENT VARIABLES
    # TIME CALCS
    startTime = datetime.datetime.combine(datetime.date.today(), times[0])
    print(startTime)
    endTime = datetime.datetime.combine(datetime.date.today(), times[-1])
    totalTime = endTime - startTime

    totalTime = [round(totalTime.total_seconds() / 3600,3)] # convert to hours

    # Get the difference between datetimes (as timedelta)
    if timeAt30 != 'N/A':
        timeAt30 = datetime.datetime.combine(datetime.date.today(), timeAt30)
        timeTo30 = timeAt30 - startTime
        timeTo30 = [round(timeTo30.total_seconds() / 3600,3)]
    else:
        timeTo30 = 'N/A'

    # MAX TEMP CALCS
    maxTemp1 = max(t1)
    maxIdx = t1.index(max(t1))
    maxTemp2 = max(t2) # max temp of control

    # MAX TIME CALCS - get time it takes to get to max temp
    maxTime = datetime.datetime.combine(datetime.date.today(), times[maxIdx])
    maxTime = maxTime - startTime
    maxTime = [round(maxTime.total_seconds() / 3600,3)]


    # find heat gain required to raise temperature to max temp
    cp = 4.184 # [kJ/kg*K] heat capacity of water
    heatGain = round(float(flow)*cp*(maxTemp1 - t1[0]),3)

    # get temp difference between initial temp and max temp
    tempDiff = round(maxTemp1 - t1[0],2)

    # getlow/high ambient temp
    minTempAmb = min(t3)
    maxTempAmb = max(t3)

    print("Max Temp 1: ", maxTemp1)
    print("Max Temp 2: ", maxTemp2)
    print("Time to Max: ", maxTime)
    print("Time to Reach 30 C: ", timeTo30)
    print("Temp Diff: ", tempDiff)
    print("Min/Max Temp Amb: ", minTempAmb, maxTempAmb)

    print("i0, Dates", dates)
    print("i1, Times",timeData)
    print("i2, T1", t1)
    print("i3, T2", t2)
    print("i4, T3", t3)
    print("i5, T4", t4)

    print(len(times),len(t1),len(t2),len(t3))
    print("Time vs. Tank Temp")

    return [startTime],tempDiff,maxTemp1,maxTemp2,timeTo30,totalTime,maxTime,heatGain,minTempAmb,maxTempAmb,fileData


[startTime],tempDiff,maxTemp1,maxTemp2,timeTo30,totalTime,maxTime,heatGain,minTempAmb,maxTempAmb,fileData = analyze("Test.txt",0.1)


# WRITING TO EXCEL
def write(weather,flow,wind,startTime,totalTime,other,tempDiff,maxTemp1,maxTemp2,timeTo30,maxTime,heatGain,minTempAmb,maxTempAmb,fileName,fileData):

    if len(weather) == 0:
        weather = ['']

    fileName = Path(fileName).stem
    # reads the column headers from the Excel sheet
    df = pd.read_excel('solarData.xlsx',sheet_name='Sheet1')

    # These are the input headers in the data entry file
    headers = ['Testing Date', 'Weather Conditions', 'Avg Flow Rate [kg/s]', 'Wind Speed [m/s]', 'Start Time', 'Test Duration [hrs]', 'Other Tested Variables', 'Max Temp Diff [C]', 'Max Temp [C]', 'Time to Reach 30C [hrs]', 'Time to Reach Max Temp [hrs]', 'Heat Gain to Reach Max Temp [kW]']

    testData = {'Testing Date': [fileName],
                'Weather Conditions': weather,
                'Avg Flow Rate [kg/s]': [flow],
                'Wind Speed [m/s]': [wind],
                'Start Time': startTime,
                'Test Duration [hrs]': totalTime,
                'Min T_a [C]': [minTempAmb],
                'Max T_a [C]': [maxTempAmb],
                'Significant Test Notes: ':[other],
                'Max Control Temp [C]':[maxTemp2],
                'Max Tank Temp [C]': [maxTemp1],
                'Max Temp Diff [C]': [tempDiff],
                'Time to Reach 30C [hrs]': [timeTo30],
                'Time to Reach Max Temp [hrs]': maxTime,
                'Heat Gain to Reach Max Temp [kW]': [heatGain]}

    df = df.append(testData, ignore_index=True)
    print(testData.values())
    print("works")

    values = list(testData.values())
    for i in range(len(values)):
        print(values[i], ":")
        print(len(values[i]))
    # new dataframe with same columns
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter('solarData.xlsx', engine='openpyxl')
    writer.book = load_workbook('solarData.xlsx')

    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    print("works")
    # read existing file
    reader = pd.read_excel(r'solarData.xlsx')
    df.to_excel(writer, sheet_name='Sheet1',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()

    # WRITING TO SECOND SHEET
    df = pd.read_excel('solarData.xlsx',sheet_name='Sheet2')
    testData = {'Date':fileName,'Time': fileData[0], 'Temp1 [C]': fileData[1], 'Temp2 [C]': fileData[2], 'Temp_a[C]': fileData[3]}
    print(fileData)
    print(fileName)
    df = df.append(testData, ignore_index=True,sort=False)
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter('solarData.xlsx', engine='openpyxl')
    writer.book = load_workbook('solarData.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(r'solarData.xlsx',sheet_name='Sheet2')
    df.to_excel(writer, sheet_name='Sheet2',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()


################
################
