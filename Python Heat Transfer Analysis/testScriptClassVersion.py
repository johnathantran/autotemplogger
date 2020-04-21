# Team Biodigester
# import modules
import datetime
import os
import pandas as pd # allows working with Excel files
import PySimpleGUI as sg # allows creation of user interface
from openpyxl import load_workbook

# defines class to contain all attributes and methods to analyze temperature data
class tempDataClass:

    # TEMPERATURE SENSOR DEFINITIONS
    # MAKE SURE THAT YOUR SENSORS MATCH THE FOLLOWING DEFINITIONS OTHERWISE THE ANALYSIS
    # RESULTS WILL BE INACCURATE

    # Temp 1: tank temp (running through solar collector)
    # Temp 2: control temp (tank of water not connected to solar collector)
    # Temp 3: ambient temp (environment temp)
    # Temp 4,5,6 ... n: whatever additional temp sensors you want to add, program will not analyze
    # these additional sensor data but will add them to the master Excel sheet on Sheet 2

    def __init__(self,**kwargs):
        
        self.kwargs = kwargs
        self.dates = kwargs['dates']
        self.times = kwargs['times']
        self.temps = [val for val in kwargs.values()][2: ] # initialize list of lists of temperature data
    
    # main method used to analyze temp data
    # INPUTS: flow rate in kg/s
    # OUTPUTS: None
    def analyze(self,flow):

        self.flow = flow
        
        # DEPENDENT VARIABLES
        for i,val in enumerate(self.times):

            # convert time data to datetime
            x = datetime.datetime.strptime(str(val), '%H:%M:%S')
            self.times[i] = x.time()

        # find time it takes the tank to reach 30 C
        for i,temp_val in enumerate(self.temps[0]):

            # assume temp never reached 30 C until data reveals otherwise
            time_at_30 = 'N/A'

            if temp_val >= 30:
                time_at_30 = self.times[i]
                break

        # TIME CALCS
        start_time = datetime.datetime.combine(datetime.date.today(), self.times[0])
        end_time = datetime.datetime.combine(datetime.date.today(), self.times[-1])
        total_time = [round((end_time - start_time).total_seconds() / 3600,3)] # convert to hours

        # Get the difference between datetimes (as timedelta)
        if time_at_30 != 'N/A':
            time_at_30 = datetime.datetime.combine(datetime.date.today(), time_at_30)
            time_to_30 = time_at_30 - start_time
            time_to_30 = [round(time_to_30.total_seconds() / 3600,3)]
        else:
            time_to_30 = 'N/A'


        # MAX TEMP CALCS
        t1,t2,t3 = self.temps[0], self.temps[1], self.temps[2]
        max_temp1 = max(t1)
        max_temp2 = max(t2) # max temp of tank, control
        maxIdx = t1.index(max(t1))

        # MAX TIME CALCS - get time it takes to get to max temp
        max_time = datetime.datetime.combine(datetime.date.today(), self.times[maxIdx])
        max_time = [round((max_time - start_time).total_seconds() / 3600,3)]

        # find heat gain required to raise temperature to max temp
        cp = 4.184 # [kJ/kg*K] heat capacity of water
        heat_gain = round(float(self.flow)*cp*(max_temp1 - t1[0]),3)

        # get temp difference between initial temp and max temp
        temp_diff = round(max_temp1 - t1[0],2)

        # getlow/high ambient temp
        min_temp_amb, max_temp_amb = min(t3), max(t3)
        
        # assign variables as class attributes
        self.temp_diff = temp_diff
        self.start_time = start_time
        self.max_temp1 = max_temp1
        self.max_temp2 = max_temp2
        self.time_to_30 = time_to_30
        self.total_time = total_time
        self.max_time = max_time
        self.heat_gain = heat_gain
        self.min_temp_amb = min_temp_amb
        self.max_temp_amb = max_temp_amb

        return


    # defines how an instance of the class is printed
    def __str__(self):

        dateStr = "Dates: " + str(self.dates) + "\n"
        timeStr = "Times: " + str(self.times) + "\n"

        tempStr = ""
        for idx, tempList in enumerate(self.temps):
            tempStr = tempStr + "Temp " + str(idx+1) + ": " + str(tempList) + "\n"

        printStr = "\nFILE DATA: \n" + dateStr + timeStr + tempStr

        return printStr

# END CLASS DEFINITION
#####################
#####################
#####################


# function used to parse input text file
# INPUTS: name of input text file
# OUTPUTS: kwargs (dictionary of variable length, contains dates, times, 'n' number of temp columns)
def parseFile(file_name):

    kwargs = {"dates": [], "times":[]} # initialize kwargs dict

    # open data file
    with open(file_name) as tsv:

        line = tsv.readline()
        while line:
            line = tsv.readline()

            # break the loop if the line is empty
            if line == '': break

            # converts line string to a list split on the "|" char
            line = [x.strip() for x in line.split('|')]
            line.pop() # takes off new line char

            for i, val in enumerate(line):

                if str(i-1) not in kwargs and i not in [0,1]:
                    kwargs[str(i-1)] = []

                if i == 0: kwargs["dates"].append(val)

                elif i == 1: kwargs["times"].append(val)

                else: kwargs[str(i-1)].append(float(val))

    return kwargs

# function used to write analysis results to Excel
# INPUTS: inputs from the GUI, and an instance of the tempClassData object
# OUTPUTS: None, but will update the master Excel file
def write(inputs,obj):

    output_file = inputs["output_file"]

    if len(inputs['weather']) == 0: weather = ['']

    #file_name = Path(file_name).stem
    # reads the column headers from the Excel sheet
    df = pd.read_excel(output_file,sheet_name='Sheet1')

    # These are the input headers in the data entry file for reference
    headers = ['Testing Date', 'Weather Conditions', 'Avg Flow Rate [kg/s]', 'Wind Speed [m/s]', 'Start Time', 'Test Duration [hrs]', 'Other Tested Variables', 'Max Temp Diff [C]', 'Max Temp [C]', 'Time to Reach 30C [hrs]', 'Time to Reach Max Temp [hrs]', 'Heat Gain to Reach Max Temp [kW]']

    testData = {'Testing Date': [obj.dates[0]],
                'Weather Conditions': inputs['weather'],
                'Avg Flow Rate [kg/s]': [obj.flow],
                'Wind Speed [m/s]': inputs['wind'],
                'Start Time': obj.start_time,
                'Test Duration [hrs]': obj.total_time,
                'Min T_a [C]': [obj.min_temp_amb],
                'Max T_a [C]': [obj.max_temp_amb],
                'Significant Test Notes: ':inputs['other'],
                'Max Control Temp [C]':[obj.max_temp2],
                'Max Tank Temp [C]': [obj.max_temp1],
                'Max Temp Diff [C]': [obj.temp_diff],
                'Time to Reach 30C [hrs]': [obj.time_to_30],
                'Time to Reach Max Temp [hrs]': obj.max_time,
                'Heat Gain to Reach Max Temp [kW]': [obj.heat_gain]}

    df = df.append(testData, ignore_index=True)
    print(testData.values())

    # new dataframe with same columns
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book = load_workbook(output_file)

    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
    # read existing file
    reader = pd.read_excel(output_file)
    df.to_excel(writer, sheet_name='Sheet1',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()

    # WRITING TO SECOND SHEET
    df = pd.read_excel(output_file,sheet_name='Sheet2')
    #testData = {'Date':self.dates,'Time': self.times, 'Temp1 [C]': self.temps[0], 'Temp2 [C]': fileData[2], 'Temp_a[C]': fileData[3]}
    testData = obj.kwargs

    df = df.append(testData, ignore_index=True,sort=False)
    df = pd.DataFrame(testData)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book = load_workbook(output_file)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(output_file,sheet_name='Sheet2')
    df.to_excel(writer, sheet_name='Sheet2',index=False,header=False,startrow=len(reader)+1)
    writer.close()
    writer.save()

################
################

# main function to run the GUI program
def main():
    
    ### GUI CODE ###
    # creates the user interface
    def_font = 'Verdana 10' # default font
    
    output_file = "solarDataTest.xlsx" # output analysis results to this Excel file
    
    # utility description
    introMain = "Select the input temperature data file from your computer. \nFill in all the required independent variable fields and click 'Analyze Data'. \n* = required field"
    
    # adjusts the size, content, and spatial arrangements of the window elements
    layout = [[sg.Text('HPD Biodigester Solar Model', font='Verdana 13 bold')],[sg.Text(introMain, font=def_font)],
    			[sg.Text('Select Input Data File*:', font=def_font, size=(20, 1))],
                [sg.InputText(size=(50, 1), font=def_font, key='file_name'),sg.FileBrowse(file_types=(("Text Files", "*.txt*"),),tooltip='Selects a data file from your computer.')],
    
                [sg.Text('Enter Environment Variables:', font='Verdana 13 bold')],
                [sg.Text('Weather: '),sg.Listbox(values=['Clear and Sunny', 'Cloudy Sunny', 'Cloudy','Light Drizzle','Rain','Heavy Rain','Sleet','Snow'], size=(20, 4),key='weather')],
                [sg.Text('Wind Speed:'),sg.InputText('',size=(5, 1), font=def_font, key='wind'),sg.Text('m/s')],
    
                [sg.Text('Enter Independent Variables:', font='Verdana 13 bold')],
                [sg.Text('Avg Flow Rate*:',font=def_font, ),sg.InputText('',size=(10, 1), font=def_font, key='flow'),sg.Text('kg/s')],
                [sg.Text('Test Notes and/or Other Variables (if any): ')],
                [sg.Multiline('',size=(30, 3), font=def_font, key='other'),],
    
                [sg.Button('Analyze Data', font=def_font, tooltip='Initial Data Analysis'),],
                [sg.Text('Analysis Output: ',font='Verdana 13 bold')],
                [sg.Multiline('',size=(50,8),disabled=True,font=def_font, key='output')],
                [sg.Text('Click Update Master Data to save your analysis data to the "solarData.xlsx" Excel file.',key='update')],
    
    			[sg.Button('Update Master Data', font=def_font, tooltip='Updates Master Excel sheet'),
                sg.Button('Clear', font=def_font, tooltip='Clears all fields.'),]]
    
    window = sg.Window('Solar Model', layout, resizable=True)
    
    while True:
        event, values = window.Read()
    
        if event is None: break
    
        # input vars
        inputs = {'output_file':output_file,
                  'weather':values['weather'],
                  'flow':(values['flow']),
                  'wind':values['wind'],
                  'other':values['other']}
        
             
        
        # event handler for 'Analyze Data' button
        if event == 'Analyze Data':
    
            # check to see if the user filled out required fields
            if values['flow'] == "": window.Element('update').Update("Please enter a flow rate.")
            
            elif values['file_name'] == "": window.Element('update').Update("Please select an input file.")
            
            else:     
                # parse the file and create a class instance
                kwargs = parseFile(values['file_name'])
                obj = tempDataClass(**kwargs)
                
                try: # call analyze method on the class object
                    obj.analyze(values['flow'])
                    window.Element('output').Update("Outside Temp (Low/High): " + str(obj.min_temp_amb) + "/" + str(obj.max_temp_amb) + " C\n"\
                                                  + "Max Temp Diff: " + str(obj.temp_diff) + " C\n"\
                                                  + "Max Temp: " + str(obj.max_temp1) + " C\n"\
                                                  + "Time to Reach 30C: " + str(obj.time_to_30) + " hrs\n"\
                                                  + "Time to Reach Max Temp: " + str(obj.max_time) + " hrs\n"\
                                                  + "Heat Gained to Reach Max Temp: " + str(obj.heat_gain) + " kW\n")
                
                    window.Element('update').Update("Your temperature data from has been successfully analyzed.")
                    
                except ValueError: window.Element('update').Update("Bro that's not even a number.")
                except IndexError: window.Element('update').Update("There was an error analyzing your file. Please check the format of your input text file.")
               

    
        # event handler for 'Update Master Data' button
        if event == 'Update Master Data':
    
            if len(values['output']) <= 1: window.Element('update').Update("Must Analyze Data before saving to Excel sheet.")
    
            else:
                try:
                    write(inputs,obj)
                    window.Element('update').Update("Your data has been successfully saved to " + output_file + ".")
                except:
                    window.Element('update').Update("Please close " + output_file + " before updating master data.")
                    
        # # event handler for 'Clear' button - clear all fields
        if event == 'Clear':
            window.Element('file_name').Update("")
            window.Element('wind').Update("")
            window.Element('other').Update("")
            window.Element('flow').Update("")
            window.Element('output').Update("")
            window.Element('update').Update("")
    
    ##########
    ##########

main()